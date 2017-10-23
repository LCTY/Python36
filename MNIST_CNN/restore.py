import os
import math
import numpy as np
import matplotlib.pyplot as plt
import tensorflow as tf
from openpyxl import Workbook
from tensorflow.examples.tutorials.mnist import input_data

mnist = input_data.read_data_sets("MNIST_data/", one_hot=True)

tf.reset_default_graph()
with tf.Session() as sess:
    save_path = "./tmp/"
    saver = tf.train.import_meta_graph(save_path + 'model.ckpt.meta')
    saver.restore(sess, tf.train.latest_checkpoint(save_path))


    # # 用於算出test資料集的精準度
    # x = tf.get_default_graph().get_tensor_by_name('x:0')
    # y_ = tf.get_default_graph().get_tensor_by_name('y_:0')
    # keep_prob = tf.get_default_graph().get_tensor_by_name('keep_prob:0')
    # y_conv = tf.get_default_graph().get_tensor_by_name('y_conv:0')
    #
    # correct_prediction = tf.equal(tf.argmax(y_conv, 1), tf.argmax(y_, 1))
    # accuracy = tf.reduce_mean(tf.cast(correct_prediction, tf.float32))
    # train_accuacy = accuracy.eval(feed_dict={x: mnist.test.images, y_: mnist.test.labels, keep_prob: 1.0})
    # print(train_accuacy)


    # # get activation value for every data set
    # x = tf.get_default_graph().get_tensor_by_name('x:0')
    # y_ = tf.get_default_graph().get_tensor_by_name('y_:0')
    # keep_prob = tf.get_default_graph().get_tensor_by_name('keep_prob:0')
    all_vars = tf.trainable_variables()
    # # pooling layer, last layer not in tf.trainable_variables()
    # y_conv = tf.get_default_graph().get_tensor_by_name('y_conv:0')
    # all_vars.append(y_conv)
    # for var in all_vars:
    #     tensor = tf.get_default_graph().get_tensor_by_name(var.name)
    #     activation_value = tensor.eval(feed_dict={x: mnist.test.images, y_: mnist.test.labels, keep_prob: 1.0})
    #     activation_value = activation_value.ravel()
    #     plt.title(var.name)
    #     plt.hist(activation_value)
    #     save_path = 'hist/'
    #     if not os.path.exists(save_path):
    #         os.makedirs(save_path)
    #     plt.savefig(save_path + 'activation_' + var.name[:var.name.find(':')] + ".png")
    #     plt.close()

    # # 用於將weight輸出成excel資料
    # wb = Workbook()
    # ws = wb.active
    #
    # all_vars = tf.trainable_variables()
    # for idx, val in enumerate(all_vars):
    #     if idx == 1:
    #         break
    #
    #     ws.cell(row=1, column=idx+1, value=val.name)
    #
    #     sv = tf.reshape(val, [-1])
    #     weight = sess.run(sv)
    #
    #     for index, value in enumerate(weight):
    #         ws.cell(row=index+2, column=idx+1, value=value)
    #
    # wb.save("sample.xlsx")


    # # 用於印出直方圖
    # all_vars = tf.trainable_variables()
    # for idx, val in enumerate(all_vars):
    #     # if idx == 1:
    #     #     break
    #     sv = tf.reshape(val, [-1])
    #     weight = sess.run(sv)
    #     for index, value in enumerate(weight):
    #         weight[index] = math.log(abs(value), 2)
    #     plt.title(val.name)
    #     plt.hist(weight, bins=range(-12, 0, 1))
    #     save_path = 'hist/'
    #     if not os.path.exists(save_path):
    #         os.makedirs(save_path)
    #     plt.savefig(save_path + str(idx) + ".png")
    #     plt.close()
